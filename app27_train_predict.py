import pandas as pd
import numpy as np
import os
import logging
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense, Input, Dropout
from tensorflow.keras.callbacks import EarlyStopping
import matplotlib.pyplot as plt
from matplotlib import rcParams
import openpyxl
from openpyxl.utils import get_column_letter

# 한글 폰트를 Malgun Gothic으로 설정
rcParams['font.family'] = 'Malgun Gothic'
rcParams['axes.unicode_minus'] = False

logging.basicConfig(level=logging.INFO)

def process_data(file_path, output_file_path):
    excel_data = pd.ExcelFile(file_path)
    processed_data = {}

    for sheet_name in excel_data.sheet_names:
        df = excel_data.parse(sheet_name)
        df.columns = df.columns.str.strip()

        # '년월일' 열을 문자열로 변환 후 날짜 처리
        if '년월일' in df.columns:
            df['년월일'] = df['년월일'].astype(str).str.replace('년', '-').str.replace('월', '-').str.replace('일', '')
            df['년월일'] = pd.to_datetime(df['년월일'], errors='coerce')
        if '예측일자' in df.columns:
            df['예측일자'] = df['예측일자'].astype(str).str.replace('년', '-').str.replace('월', '-').str.replace('일', '')
            df['예측일자'] = pd.to_datetime(df['예측일자'], errors='coerce')

        numeric_cols = ['유입량(㎥/일)', '인구수(명)', '강수량(mm)']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        if '년월' not in df.columns and '년월일' in df.columns:
            df['년월'] = df['년월일'].dt.to_period('M')

        if '인구수(명)' in df.columns:
            growth_rate = 0.02
            df['인구수(명)'] *= (1 + growth_rate)

        if '강수량(mm)' in df.columns and '년월일' in df.columns:
            df['월'] = df['년월일'].dt.month
            df['일'] = df['년월일'].dt.day
            df['강수량(mm)'] = df.groupby(['월', '일'])['강수량(mm)'].transform(lambda x: x.fillna(x.mean()))

        df.drop(columns=['월', '일'], inplace=True)
        processed_data[sheet_name] = df

    with pd.ExcelWriter(output_file_path) as writer:
        for sheet_name, df in processed_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def create_and_train_lstm_model(X_train, y_train, time_steps):
    logging.info("LSTM 모델을 생성하고 훈련합니다.")
    
    model = Sequential()
    model.add(Input(shape=(time_steps, X_train.shape[2])))
    model.add(LSTM(50, return_sequences=True))
    model.add(Dropout(0.2))
    model.add(LSTM(50))
    model.add(Dropout(0.2))
    model.add(Dense(1))
    model.compile(optimizer='adam', loss='mean_squared_error')

    early_stopping = EarlyStopping(monitor='loss', patience=10)
    model.fit(X_train, y_train, epochs=100, batch_size=32, verbose=1, callbacks=[early_stopping])
    
    return model

def prepare_data_for_lstm(df, time_steps, future_days):
    features = ['강수량(mm)', '인구수(명)']
    target = '유입량(㎥/일)'
    
    feature_data = df[features].values
    target_data = df[target].values.reshape(-1, 1)
    
    feature_scaler = MinMaxScaler()
    target_scaler = MinMaxScaler()
    
    scaled_features = feature_scaler.fit_transform(feature_data)
    scaled_target = target_scaler.fit_transform(target_data)
    
    X = []
    y = []
    for i in range(time_steps, len(scaled_features)):
        X.append(scaled_features[i-time_steps:i, :])
        y.append(scaled_target[i])
    X, y = np.array(X), np.array(y)
    
    return X, y, feature_scaler, target_scaler

def predict_existing_data(df, model, feature_scaler, target_scaler, time_steps):
    logging.info("기존 데이터를 기반으로 예측합니다.")
    features = ['강수량(mm)', '인구수(명)']
    target = '유입량(㎥/일)'

    feature_data = df[features].values
    target_data = df[target].values.reshape(-1, 1)

    scaled_features = feature_scaler.transform(feature_data)
    scaled_target = target_scaler.transform(target_data)

    X = []
    for i in range(time_steps, len(scaled_features)):
        X.append(scaled_features[i-time_steps:i, :])
    X = np.array(X)

    predictions_scaled = model.predict(X)
    predictions = target_scaler.inverse_transform(predictions_scaled)

    df['예측 유입량(㎥/일)'] = np.nan
    df.loc[time_steps:, '예측 유입량(㎥/일)'] = predictions.flatten()
    df['예측일자'] = df['년월일']

    # 예측 유입량 NaN 값 처리: 이전 값으로 채움
    df['예측 유입량(㎥/일)'] = df['예측 유입량(㎥/일)'].ffill()

    # 예측값 출력
    print("기존 데이터 예측 유입량:", df['예측 유입량(㎥/일)'].to_list())

    return df

def predict_future_values(df, model, feature_scaler, target_scaler, time_steps, future_days, avg_rainfall_per_day):
    logging.info("미래의 예측 유입량을 계산합니다.")
    
    features = ['강수량(mm)', '인구수(명)']
    
    current_features = df[features].values[-time_steps:]
    predictions_list = []
    
    current_date = df['년월일'].max()
    
    if pd.isna(current_date):
        logging.error("현재 날짜가 유효하지 않습니다.")
        return pd.DataFrame()  # 빈 데이터프레임 반환

    business_code = df['사업소코드'].iloc[0]
    business_name = df['사업소'].iloc[0]
    
    last_predicted_value = df['유입량(㎥/일)'].mean()  # 기본값으로 평균값 설정

    for day in range(future_days):
        scaled_features = feature_scaler.transform(current_features)
        X_future = np.array([scaled_features])
        
        future_predictions_scaled = model.predict(X_future)
        future_predictions = target_scaler.inverse_transform(future_predictions_scaled)
        
        print(f"예측 입력 데이터: {scaled_features}")  # 입력 데이터 출력
        print(f"예측된 스케일된 값: {future_predictions_scaled}")  # 스케일된 예측값 출력
        print(f"예측된 값: {future_predictions}")  # 역변환된 예측값 출력
        
        current_date += pd.DateOffset(days=1)
        
        avg_rainfall = avg_rainfall_per_day.get(current_date.day, 0)
        
        new_row = np.array([avg_rainfall, df['인구수(명)'].iloc[-1]])
        
        current_features = np.vstack([current_features[1:], new_row])
        
        predicted_value = future_predictions[0][0]

        # 예측값이 NaN이면 마지막 예측값으로 대체
        if pd.isna(predicted_value):
            predicted_value = last_predicted_value  # 마지막 예측값으로 대체

        predictions_list.append({
            '년월일': current_date, 
            '예측 유입량(㎥/일)': predicted_value,
            '사업소코드': business_code, 
            '사업소': business_name,
            '강수량(mm)': avg_rainfall,
            '인구수(명)': df['인구수(명)'].iloc[-1],
            '년월': current_date.to_period('M')  # 추가된 부분
        })

        # 마지막 예측값 업데이트
        last_predicted_value = predicted_value
    
    future_predictions_df = pd.DataFrame(predictions_list)

    # 예측 유입량 NaN 값 처리: 이전 값으로 채움
    future_predictions_df['예측 유입량(㎥/일)'] = future_predictions_df['예측 유입량(㎥/일)'].ffill()

    # 예측값 출력
    print("미래 데이터 예측 유입량:", future_predictions_df['예측 유입량(㎥/일)'].to_list())

    future_predictions_df['예측일자'] = future_predictions_df['년월일']
    
    return future_predictions_df

def calculate_average_rainfall(df):
    logging.info("같은 날짜의 강수량 평균을 계산합니다.")
    
    df['날짜'] = pd.to_datetime(df['년월일'])
    avg_rainfall_per_day = df.groupby(df['날짜'].dt.day)['강수량(mm)'].mean().to_dict()
    
    return avg_rainfall_per_day

def save_predictions_to_excel(df, output_path, sheet_name):
    logging.info(f"{sheet_name} 시트의 예측 결과를 Excel 파일에 저장합니다.")
    
    if not os.path.exists(output_path):
        df.to_excel(output_path, index=False, sheet_name=sheet_name)
    else:
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            worksheet = writer.sheets[sheet_name]
            for column in df.columns:
                column_width = max(df[column].astype(str).map(len).max(), len(column))
                col_letter = get_column_letter(df.columns.get_loc(column) + 1)
                worksheet.column_dimensions[col_letter].width = column_width

def save_performance_metrics(metrics_output_path, metrics):
    logging.info(f"모델 성능 지표를 Excel 파일에 저장합니다.")
    metrics_df = pd.DataFrame(metrics)
    
    if not os.path.exists(metrics_output_path):
        metrics_df.to_excel(metrics_output_path, index=False, sheet_name='성능 지표')
    else:
        with pd.ExcelWriter(metrics_output_path, engine='openpyxl', mode='a') as writer:
            if '성능 지표' in writer.book.sheetnames:
                del writer.book['성능 지표']
            metrics_df.to_excel(writer, index=False, sheet_name='성능 지표')

def plot_and_save_graphs(df_existing_predictions, df_future_predictions, output_graph_dir, sheet_name):
    plt.figure(figsize=(14, 7))
    plt.plot(df_existing_predictions['년월일'], df_existing_predictions['유입량(㎥/일)'], label='Actual Inflow')
    plt.plot(df_existing_predictions['년월일'], df_existing_predictions['예측 유입량(㎥/일)'], label='Predicted Inflow')
    
    if df_future_predictions is not None:
        plt.plot(df_future_predictions['년월일'], df_future_predictions['예측 유입량(㎥/일)'], label='Future Predictions', linestyle='--')

    plt.xlabel('Date')
    plt.ylabel('Inflow (㎥/일)')
    plt.title(f'{sheet_name} - Inflow Prediction vs Actual')
    plt.legend()
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    graph_path = os.path.join(output_graph_dir, f'{sheet_name}_예측_그래프.png')
    plt.savefig(graph_path, format='png')
    plt.close()

def evaluate_model_performance(df, predicted_column, actual_column, target_scaler, sheet_name):
    logging.info("모델 성능을 평가합니다.")
    y_true = df[actual_column].dropna()
    y_pred = df[predicted_column].dropna()

    common_index = y_true.index.intersection(y_pred.index)
    y_true = y_true.loc[common_index]
    y_pred = y_pred.loc[common_index]

    mse = mean_squared_error(y_true, y_pred)
    mae = mean_absolute_error(y_true, y_pred)
    rmse = np.sqrt(mse)
    r_squared = r2_score(y_true, y_pred)
    
    metrics = {
        'Sheet Name': [sheet_name],
        'MSE': [mse],
        'MAE': [mae],
        'RMSE': [rmse],
        'R²': [r_squared]
    }

    return metrics

def process_sheets_for_lstm(file_path, output_path, model_dir, output_graph_dir, metrics_output_path):
    all_sheets = pd.ExcelFile(file_path).sheet_names
    all_metrics = []

    for sheet_name in all_sheets:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # '년월일' 열 확인 및 변환
        if '년월일' in df.columns:
            df['년월일'] = pd.to_datetime(df['년월일'], format='%Y-%m-%d %I:%M:%S %p', errors='coerce')
            if df['년월일'].isnull().all():
                logging.warning(f"'{sheet_name}' 시트에 유효한 '년월일' 값이 없습니다. 스킵합니다.")
                continue
        else:
            logging.warning(f"'{sheet_name}' 시트에 '년월일' 열이 없습니다. 스킵합니다.")
            continue

        df.sort_values('년월일', inplace=True)

        # 유입량(㎥/일) 컬럼의 결측치를 0으로 채움
        if '유입량(㎥/일)' in df.columns:
            df['유입량(㎥/일)'] = df['유입량(㎥/일)'].fillna(0)

        # 모델 훈련 및 예측 로직
        time_steps = 30  
        future_days = 30  
        X, y, feature_scaler, target_scaler = prepare_data_for_lstm(df, time_steps, future_days)

        model = create_and_train_lstm_model(X, y, time_steps)

        model_save_path = os.path.join(model_dir, f'{sheet_name}_lstm_model.keras')
        model.save(model_save_path)
        logging.info(f"'{sheet_name}' 시트의 모델을 저장했습니다: {model_save_path}")

        df_existing_predictions = predict_existing_data(df.copy(), model, feature_scaler, target_scaler, time_steps)

        # 예측 유입량 NaN 처리
        df_existing_predictions['예측 유입량(㎥/일)'] = df_existing_predictions['예측 유입량(㎥/일)'].ffill().bfill().fillna(0)

        # 예측값이 NaN인 경우 처리
        assert not df_existing_predictions['예측 유입량(㎥/일)'].isnull().any(), f"{sheet_name} 시트에서 '예측 유입량(㎥/일)'에 결측치가 남아 있습니다."

        avg_rainfall_per_day = calculate_average_rainfall(df)
        df_future_predictions = predict_future_values(df.copy(), model, feature_scaler, target_scaler, time_steps, future_days, avg_rainfall_per_day)

        df_combined_predictions = pd.concat([df_existing_predictions, df_future_predictions], ignore_index=True)

        # 예측 유입량(㎥/일) 열의 결측치를 처리: 앞뒤 값으로 채우고, 남은 결측치를 0으로 채움
        df_combined_predictions['예측 유입량(㎥/일)'] = df_combined_predictions['예측 유입량(㎥/일)'].ffill().bfill().fillna(0)

        # 결과 Excel에 저장
        save_predictions_to_excel(df_combined_predictions, output_path, f'{sheet_name}')

        metrics = evaluate_model_performance(df_existing_predictions, '예측 유입량(㎥/일)', '유입량(㎥/일)', target_scaler, sheet_name)
        all_metrics.append(metrics)

        plot_and_save_graphs(df_existing_predictions, df_future_predictions, output_graph_dir, sheet_name)

    save_performance_metrics(metrics_output_path, all_metrics)

if __name__ == '__main__':
    file_path = '예측된_수정된_취합_일자_강수량_인구_유입량_20240829_제공.xlsx'
    output_path = '예측결과.xlsx'
    model_dir = 'models'
    output_graph_dir = 'graphs'
    metrics_output_path = 'metrics.xlsx'

    # 데이터 전처리 실행
    process_data(file_path, '처리된_데이터.xlsx')

    # LSTM 모델 처리 실행
    process_sheets_for_lstm('처리된_데이터.xlsx', output_path, model_dir, output_graph_dir, metrics_output_path)
